from time import sleep
import pdfplumber
import pandas as pd
import re
from typing import List, Dict, Optional
import yaml
import os
import inquirer
from datetime import datetime
import sys
import readchar

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

class PDFExtractor:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        
    def extract_text(self) -> str:
        text_data = []
        with pdfplumber.open(self.pdf_path) as pdf:
            for page in pdf.pages:
                text_data.append(page.extract_text())
        return "\n".join(text_data)

class TextParser:
    def __init__(self, title_keywords: List[str]):
        self.title_keywords = title_keywords
        self.title_pattern = re.compile(rf"^({'|'.join(title_keywords)})[\w\s/-]*")
        # Improved date pattern to better capture German date formats
        self.date_pattern = re.compile(r"(Jan|Feb|Mrz|Apr|Mai|Jun|Jul|Aug|Sep|Okt|Nov|Dez)\s*(\d{2})(?:\s*-\s*(Jan|Feb|Mrz|Apr|Mai|Jun|Jul|Aug|Sep|Okt|Nov|Dez)\s*(\d{2}))?")
    
    def parse_sections(self, text: str) -> Dict[str, List[str]]:
        sections = {}
        current_title = None
        
        for line in text.split("\n"):
            line = line.strip()
            if self.title_pattern.match(line):
                current_title = line
                sections[current_title] = []
            elif current_title and line and not line.startswith("Ernte"):
                sections[current_title].append(line)
        return sections

class DataProcessor:
    def __init__(self, date_pattern):
        self.date_pattern = date_pattern
        self.month_order = {
            'Jan': 1, 'Feb': 2, 'Mrz': 3, 'Apr': 4, 'Mai': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Okt': 10, 'Nov': 11, 'Dez': 12
        }
        
    def parse_date_period(self, match):
        """Enhanced date period parser"""
        if not match:
            return None
            
        groups = match[0]
        if len(groups) >= 2:  # At least one date
            start_month, start_year = groups[0], groups[1]
            
            if len(groups) >= 4 and groups[2] and groups[3]:  # Date range
                end_month, end_year = groups[2], groups[3]
                return f"{start_month} {start_year} - {end_month} {end_year}"
            else:  # Single date
                return f"{start_month} {start_year}"
        return None

    def sort_periods(self, periods):
        """Sort periods chronologically"""
        def period_key(period):
            if ' - ' in period:
                start, _ = period.split(' - ')
                month, year = start.split()
            else:
                month, year = period.split()
            return int(year), self.month_order[month]
            
        return sorted(periods, key=period_key)
        
    def validate_price(self, price: float) -> bool:
        """Enhanced price validation"""
        return (
            isinstance(price, (int, float)) and 
            100 <= price <= 1000 and  # Basic range check
            price % 1 <= 0.99  # Ensure reasonable decimal places
        )
        
    def calculate_statistics(self, prices: List[float]) -> Dict:
        """Calculate price statistics"""
        valid_prices = [p for p in prices if self.validate_price(p)]
        if not valid_prices:
            return {"avg": "-", "min": "-", "max": "-", "count": 0}
            
        return {
            "avg": round(sum(valid_prices) / len(valid_prices), 2),
            "min": round(min(valid_prices), 2),
            "max": round(max(valid_prices), 2),
            "count": len(valid_prices)
        }
        
    def process_data(self, sections: Dict[str, List[str]], file_index: int) -> dict:
        all_data = {}
        
        for title, lines in sections.items():
            if title not in all_data:
                all_data[title] = {}
                
            for line in lines:
                dates = self.date_pattern.findall(line)
                if not dates:
                    continue
                    
                period = self.parse_date_period(dates)
                if not period:
                    continue
                    
                # Extract prices with improved regex
                prices = [
                    p for p in re.findall(r'\d{3}(?:,\d{1,2})?', line)
                    if "Prämie" not in line
                ]
                
                if not prices:
                    continue
                    
                if period not in all_data[title]:
                    all_data[title][period] = {
                        "prices": [],
                        "file_averages": {},
                        "raw_prices": []
                    }
                
                try:
                    numeric_prices = [float(p.replace(",", ".")) for p in prices]
                    valid_prices = [p for p in numeric_prices if self.validate_price(p)]
                    
                    if valid_prices:
                        all_data[title][period]["prices"].extend(valid_prices)
                        all_data[title][period]["raw_prices"].extend(prices)
                        stats = self.calculate_statistics(valid_prices)
                        all_data[title][period]["file_averages"][file_index] = stats["avg"]
                except ValueError:
                    continue
                    
        return all_data

class ExcelExporter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        
    def sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize sheet name to comply with Excel restrictions:
        - Maximum 31 characters
        - No special characters: [ ] : * ? / \
        - Cannot be empty or consist only of spaces
        """
        # Remove or replace invalid characters
        invalid_chars = r'[\[\]:*?/\\]'  # Changed from r'[[\]:*?/\\]'
        name = re.sub(invalid_chars, '_', name)
        
        # Trim spaces from start and end
        name = name.strip()
        
        # Ensure name is not empty
        if not name:
            name = "Sheet"
            
        # Truncate to 31 characters if needed
        if len(name) > 31:
            name = name[:31]
            
        return name
        
    def export_to_excel(self, combined_data: dict, num_files: int, file_names: List[str]):
        if not combined_data:
            print("Warning: No data found to process")
            return
            
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl')
        
        used_names = set()
        
        # First, create main sheets
        for title, periods in combined_data.items():
            base_name = self.sanitize_sheet_name(title)
            sheet_name = base_name
            counter = 1
            
            while sheet_name in used_names:
                sheet_name = f"{base_name}_{counter}"
                counter += 1
            
            used_names.add(sheet_name)
            
            processor = DataProcessor(None)
            sorted_periods = processor.sort_periods(periods.keys())
            
            # Create new data structure for horizontal layout
            headers = ["Data Type"] + sorted_periods
            rows = []
            
            # Add rows for each file using actual file names
            for i in range(num_files):
                row = [file_names[i]]  # Use actual filename instead of "File N"
                for period in sorted_periods:
                    value = periods[period]["file_averages"].get(i, "-")
                    row.append(value)
                rows.append(row)
            
            # Create DataFrame with the new structure
            df = pd.DataFrame(rows, columns=headers)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply formatting as before
            worksheet = writer.sheets[sheet_name]
            self.apply_formatting(worksheet)

        # Create details sheet
        details_sheet = writer.book.create_sheet("Details")
        details_headers = ["File Name", "Type", "Price"]
        details_rows = []

        # Collect all file details with actual file names
        for title, periods in combined_data.items():
            for period, data in periods.items():
                for file_idx, price in data["file_averages"].items():
                    if price != "-":
                        details_rows.append([
                            file_names[file_idx],  # Use actual filename
                            f"{title} {period}",
                            price
                        ])

        # Create and format details sheet
        details_df = pd.DataFrame(details_rows, columns=details_headers)
        details_df.to_excel(writer, sheet_name="Details", index=False)
        
        # Format details sheet
        details_worksheet = writer.sheets["Details"]
        self.apply_formatting(details_worksheet)
        
        # Adjust column widths for details sheet
        for column in details_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            details_worksheet.column_dimensions[column_letter].width = max_length + 4

        writer.close()

    def apply_formatting(self, worksheet):
        """Helper method for common formatting"""
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if cell.value is None or cell.value == "":
                    cell.value = "-"
        
        # Freeze first row
        worksheet.freeze_panes = "A2"
        
        # Remove grid lines
        worksheet.sheet_view.showGridLines = False

class Settings:
    def __init__(self):
        self.settings_path = get_resource_path("settings.yaml")
        self.default_settings = {
            "title_keywords": ["Weizen", "F-Weizen"]
        }

    def load_settings(self):
        if os.path.exists(self.settings_path):
            with open(self.settings_path, 'r') as f:
                return yaml.safe_load(f)
        else:
            with open(self.settings_path, 'w') as f:
                yaml.dump(self.default_settings, f)
            return self.default_settings

    def save_settings(self, settings):
        with open(self.settings_path, 'w') as f:
            yaml.dump(settings, f)


def clear_console():
    if os.name == 'nt':  # for Windows
        os.system('cls')
    else:  # for Unix/Linux/MacOS
        os.system('clear')

def greet_user():
    print("\nPDF Helper v1.0")
    print("Author: Ocean Script Studio")
    print("-" * 30)
    print()

def retry_prompt(prompt_func, max_retries=3):
    """Retry a prompt function if it fails."""
    for attempt in range(max_retries):
        try:
            return prompt_func()
        except Exception as e:
            print(f"Error: {e}. Retrying ({attempt + 1}/{max_retries})...")
            sleep(1)
    print("Max retries reached. Exiting.")
    exit()

class MenuHandler:
    def __init__(self):
        self.settings_manager = Settings()
        self.settings = self.settings_manager.load_settings()

    def show_main_menu(self):
        clear_console()
        greet_user()
        questions = [
            inquirer.List('choice',
                 message="Select an option:",
                 choices=[
                     ('Process PDF files', '1'),
                     ('Settings', '2'),
                     ('Exit', '3')
                 ])
        ]
        return retry_prompt(lambda: inquirer.prompt(questions)['choice'])

    def show_settings_menu(self):
        clear_console()
        greet_user()
        print("\nCurrent title keywords:", self.settings["title_keywords"])
        keyword_questions = [
            inquirer.List('action',
                     message="Settings:",
                     choices=[
                     ('Keep current keywords', 'keep'),
                     ('Edit keywords', 'edit')
                     ])
        ]
        return retry_prompt(lambda: inquirer.prompt(keyword_questions)['action'])

    def edit_keywords(self):
        clear_console()
        greet_user()
        new_keywords = []
        while True:
            add_question = [
                inquirer.Text('keyword', message="Enter a keyword (or press Enter to finish)")
            ]
            keyword = retry_prompt(lambda: inquirer.prompt(add_question)['keyword'])
            if not keyword:
                break
            new_keywords.append(keyword)
        return new_keywords

class PDFFileHandler:
    def __init__(self):
        self.pdf_dir = get_resource_path("pdf-files")

    def ensure_pdf_directory(self):
        if not os.path.exists(self.pdf_dir):
            os.makedirs(self.pdf_dir)

    def get_pdf_files(self):
        pdf_files = []
        for file in os.listdir(self.pdf_dir):
            if file.lower().endswith('.pdf'):
                full_path = os.path.join(self.pdf_dir, file)
                pdf_files.append((full_path, os.path.getmtime(full_path)))
        return sorted(pdf_files, key=lambda x: x[1], reverse=True)

    def select_files(self, pdf_files):
        clear_console()
        greet_user()
        # Simplify the choices structure - just use filename and full path
        choices = [(os.path.basename(f[0]), f[0]) for f in pdf_files]
        questions = [
            inquirer.Checkbox('files',
                         message="Select PDF files to process (use space to select, enter to confirm)",
                         choices=choices)
        ]
        result = retry_prompt(lambda: inquirer.prompt(questions))
        if result is None:
            return []
            
        selected_files = []
        for file_path in result['files']:
            selected_files.append((file_path, os.path.basename(file_path)))
            
        return selected_files

def main():
    menu_handler = MenuHandler()
    pdf_handler = PDFFileHandler()

    while True:
        choice = menu_handler.show_main_menu()
        
        if choice == '1':
            pdf_handler.ensure_pdf_directory()
            pdf_files = pdf_handler.get_pdf_files()

            if not pdf_files:
                print("\nNo PDF files found in the pdf-files directory")
                continue

            print(f"\nFound {len(pdf_files)} PDF files:")
            for file, _ in pdf_files:
                print(f"- {os.path.basename(file)}")

            selected_files = pdf_handler.select_files(pdf_files)

            if 'menu' in selected_files or not selected_files:
                continue

            title_keywords = menu_handler.settings["title_keywords"]
            
            # Create excel-output directory if it doesn't exist
            if not os.path.exists(get_resource_path("excel-output")):
                os.makedirs(get_resource_path("excel-output"))
                
            # Get current date in a readable format
            current_time = datetime.now().strftime("%d-%B-%Y_%H-%M")
            
            # Initialize combined data dictionary
            combined_data = {}
            file_count = len(selected_files)
            file_names = [f[1] for f in selected_files]  # Extract filenames
            
            # Process each selected file
            for file_index, (pdf_path, filename) in enumerate(selected_files):
                pdf_extractor = PDFExtractor(pdf_path)
                pdf_text = pdf_extractor.extract_text()
                
                text_parser = TextParser(title_keywords)
                sections = text_parser.parse_sections(pdf_text)

                data_processor = DataProcessor(text_parser.date_pattern)
                file_data = data_processor.process_data(sections, file_index)
                
                # Combine data from current file
                for title, periods in file_data.items():
                    if title not in combined_data:
                        combined_data[title] = {}
                    for period, data in periods.items():
                        if period not in combined_data[title]:
                            combined_data[title][period] = {"prices": [], "file_averages": {}}
                        combined_data[title][period]["prices"].extend(data["prices"])
                        combined_data[title][period]["file_averages"].update(data["file_averages"])
                
                print(f"\nProcessed {os.path.basename(pdf_path)}")

            output_file = os.path.join(get_resource_path("excel-output"), f"Combined_Market_Data_{current_time}.xlsx")
            
            # Export combined data to Excel
            excel_exporter = ExcelExporter(output_file)
            excel_exporter.export_to_excel(combined_data, file_count, file_names)
            print(f"\nCreated combined Excel file: {output_file}")

        elif choice == '2':
            if menu_handler.show_settings_menu() == 'edit':
                new_keywords = menu_handler.edit_keywords()
                if new_keywords:
                    menu_handler.settings["title_keywords"] = new_keywords
                    menu_handler.settings_manager.save_settings(menu_handler.settings)

        elif choice == '3':
            print("\nGoodbye!")
            exit()

if __name__ == "__main__":
    main()